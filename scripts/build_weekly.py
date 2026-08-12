"""Convert the five weekly billing exports into the console import CSVs.

Usage:
    python scripts/build_weekly.py exports/ out/

Files are identified by their headers, never by filename. Raw exports and
produced CSVs must stay outside git; .gitignore already excludes exports/ and
this pipeline also excludes out/.
"""
from __future__ import annotations

import argparse
import csv
import math
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

SCHEMA = {
    "providers": ["name", "am", "exposure", "credit", "terms", "conf"],
    "customers": ["name", "am", "exposure", "credit", "terms", "conf"],
    "carriers": ["name", "role", "am", "dur", "calls", "rev", "profit", "due", "netbal", "curexp", "exp"],
    "routes": ["id", "destination", "provider", "trunk", "lcr", "buy", "sell", "profit", "profit_pct", "asr", "acd", "calls", "dur"],
    "customer_routes": ["id", "customer", "destination", "provider", "sell", "buy", "profit", "profit_pct", "asr", "acd", "calls", "dur", "rev", "exp"],
    "am_breakdown": ["id", "am", "side", "name", "profit", "calls", "dur", "rev", "exp", "routes", "sec"],
    "am_totals": ["am", "total_carriers"],
    "daily_carrier": ["id", "carrier", "day", "cust_rev", "cust_profit", "cust_dur", "cust_calls", "prov_exp", "prov_profit", "prov_dur", "prov_calls"],
}

DROP_TOKENS = {
    "tec", "com", "loc", "ltd", "limited", "llc", "pte", "srl", "inc", "co", "sal",
    "offshore", "telecom", "communications", "networks", "global", "group", "solutions",
    "carrier", "services", "international", "technologies", "holdings",
}


def clean_header(v):
    return re.sub(r"\s+", " ", str(v or "").strip())


def norm_name(v):
    s = re.sub(r"[^a-z0-9]+", " ", str(v or "").lower()).strip()
    toks = [x for x in s.split() if x not in DROP_TOKENS]
    return "".join(toks)


def identity_match(a, b):
    ka, kb = norm_name(a), norm_name(b)
    if not ka or not kb:
        return False
    if ka == kb:
        return True
    return ka.startswith(kb[:4]) or kb.startswith(ka[:4]) if min(len(ka), len(kb)) >= 4 else False


def to_float(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("%", "")
    if s.upper() in {"NULL", "NA", "N/A", "-", "--"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def first(row, *names):
    for n in names:
        if n in row and str(row[n] or "").strip() != "":
            return row[n]
    return None


def find_col(headers, exact=(), contains=()):
    hmap = {clean_header(h).lower(): clean_header(h) for h in headers}
    for x in exact:
        if x.lower() in hmap:
            return hmap[x.lower()]
    for h in headers:
        lh = clean_header(h).lower()
        if all(x.lower() in lh for x in contains):
            return clean_header(h)
    return None


def read_csv_file(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.DictReader(fh, dialect=dialect))
        headers = [clean_header(x) for x in (rows[0].keys() if rows else [])]
        return headers, [{clean_header(k): v for k, v in r.items()} for r in rows]


def read_xlsx(path, header_row=1):
    if load_workbook is None:
        raise SystemExit("openpyxl is required to read Excel exports; install it with: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    vals = list(ws.iter_rows(values_only=True))
    if len(vals) < header_row:
        return [], []
    headers = [clean_header(x) for x in vals[header_row - 1]]
    rows = []
    for raw in vals[header_row:]:
        if not any(x not in (None, "") for x in raw):
            continue
        rows.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers)) if headers[i]})
    return headers, rows


def read_any(path, lcr=False):
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return read_xlsx(path, 2 if lcr else 1)
    return read_csv_file(path)


def classify(folder):
    found = {}
    candidates = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {".csv", ".xlsx", ".xlsm", ".xls", ".xltx", ".xltm"}]
    for path in candidates:
        try:
            headers, _ = read_any(path)
        except Exception:
            continue
        hs = {clean_header(x) for x in headers}
        if {"Date", "Month", "Customer", "Provider", "Duration (m)", "Revenue", "Expense", "Profit"} <= hs:
            kind = "gross_profit"
        elif {"Customer", "Provider", "Destination", "IG ASR (%)", "ACD (min)", "PDD (ms)"} <= hs:
            kind = "full_info"
        elif {"Carrier Name", "Carrier Account Manager"} <= hs:
            kind = "carrier_details"
        else:
            # Exposure is deliberately checked for its semicolon-style headers.
            if "Carrier" in hs and "Current Exposure" in hs and "Credit Limit (Customer)" in hs:
                kind = "exposure"
            else:
                # LCR has its header on row 2, so inspect it separately.
                try:
                    h2, _ = read_any(path, lcr=True)
                except Exception:
                    h2 = []
                triple = sum(1 for h in h2 if h.lower() in {"provider - epg", "volume", "rate"})
                kind = "lcr" if triple >= 3 else None
        if kind:
            if kind in found:
                raise SystemExit(f"Multiple files match {kind}: {found[kind].name} and {path.name}")
            found[kind] = path
    required = {"gross_profit", "full_info", "lcr", "exposure", "carrier_details"}
    missing = required - found.keys()
    if missing:
        raise SystemExit("Could not identify export(s) by header: " + ", ".join(sorted(missing)))
    return found


def numeric(row, aliases):
    return to_float(first(row, *aliases)) or 0.0


def parse_day(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(s.replace("Z", "")).date().isoformat()
    except ValueError:
        return s[:10]


def parse_lcr(path):
    headers, rows = read_any(path, lcr=True)
    dest_col = find_col(headers, exact=("Destination", "Dest"), contains=("destination",)) or headers[0]
    entries = defaultdict(list)
    starts = []
    for i, h in enumerate(headers):
        if h.lower() == "provider - epg":
            starts.append(i)
    for row in rows:
        dest = str(row.get(dest_col) or "").strip()
        if not dest:
            continue
        for i in starts:
            if i + 2 >= len(headers):
                continue
            pe = str(row.get(headers[i]) or "").strip()
            vol = numeric(row, [headers[i + 1]])
            rate = to_float(row.get(headers[i + 2]))
            if not pe or rate is None:
                continue
            # LAST hyphen is the separator. EPG names may contain spaces.
            if "-" not in pe:
                continue
            provider, epg = pe.rsplit("-", 1)
            provider, epg = provider.strip(), epg.strip()
            if not provider:
                continue
            entries[(norm_name(dest), norm_name(provider))].append((dest, provider, epg, rate, vol))
    return entries


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("exports", type=Path)
    ap.add_argument("out", type=Path)
    args = ap.parse_args()
    if not args.exports.is_dir():
        raise SystemExit(f"No such exports directory: {args.exports}")
    args.out.mkdir(parents=True, exist_ok=True)
    files = classify(args.exports)

    gp_h, gp_rows = read_any(files["gross_profit"])
    fi_h, fi_rows = read_any(files["full_info"])
    ex_h, ex_rows = read_any(files["exposure"])
    cd_h, cd_rows = read_any(files["carrier_details"])
    lcr = parse_lcr(files["lcr"])

    # Gross Profit: drop ONLY rows with null Date. Provider-null rows are real.
    gp = []
    for r in gp_rows:
        if first(r, "Date") in (None, ""):
            continue
        gp.append({
            "date": parse_day(first(r, "Date")),
            "customer": str(first(r, "Customer") or "").strip(),
            "provider": str(first(r, "Provider") or "").strip() or "(unspecified)",
            "dur": numeric(r, ["Duration (m)", "Duration"]),
            "rev": numeric(r, ["Revenue"]),
            "exp": numeric(r, ["Expense"]),
            "profit": numeric(r, ["Profit"]),
        })
    gp_total = {k: sum(x[k] for x in gp) for k in ("dur", "rev", "exp", "profit")}

    # Full info destination minutes/traffic. Aliases make the converter tolerant
    # of the report's extra columns while the six distinguishing headers remain strict.
    fi = defaultdict(lambda: defaultdict(float))
    for r in fi_rows:
        customer = str(first(r, "Customer") or "").strip()
        provider = str(first(r, "Provider") or "").strip() or "(unspecified)"
        dest = str(first(r, "Destination") or "").strip() or "(unspecified)"
        key = (norm_name(customer), norm_name(provider), norm_name(dest))
        fi[key]["dur"] += numeric(r, ["Duration (m)", "Duration", "Minutes", "Duration Min"])
        fi[key]["calls"] += numeric(r, ["Calls", "Call Count", "Attempts"])
        fi[key]["asr"] = numeric(r, ["IG ASR (%)", "ASR (%)", "ASR"])
        fi[key]["acd"] = numeric(r, ["ACD (min)", "ACD", "Average Call Duration"])

    pair_dests = defaultdict(list)
    for (c, p, d), vals in fi.items():
        pair_dests[(c, p)].append((d, vals["dur"], vals))

    cr = defaultdict(lambda: {"rev": 0.0, "exp": 0.0, "profit": 0.0, "dur": 0.0, "calls": 0.0})
    daily = defaultdict(lambda: {"cust_rev": 0.0, "cust_profit": 0.0, "cust_dur": 0.0, "cust_calls": 0.0, "prov_exp": 0.0, "prov_profit": 0.0, "prov_dur": 0.0, "prov_calls": 0.0})

    for g in gp:
        pair = (norm_name(g["customer"]), norm_name(g["provider"]))
        dests = pair_dests.get(pair, [])
        if not dests or sum(max(0.0, x[1]) for x in dests) <= 0:
            dests = [(norm_name("(unspecified)"), 1.0, {"dur": 0.0, "calls": 0.0, "asr": 0.0, "acd": 0.0})]
        total_m = sum(max(0.0, x[1]) for x in dests)
        for dkey, minutes, metrics in dests:
            share = max(0.0, minutes) / total_m if total_m else 1.0 / len(dests)
            # Recover display names from the Full info row set later; normalized keys
            # are used for grouping only.
            k = (pair[0], dkey, pair[1])
            cr[k]["rev"] += g["rev"] * share
            cr[k]["exp"] += g["exp"] * share
            cr[k]["profit"] += g["profit"] * share
            # Traffic metrics come from Full info, not from the monetary split.
            cr[k]["dur"] = metrics["dur"]
            cr[k]["calls"] = metrics["calls"]
        if g["date"]:
            dk = g["date"]
            cd = g["customer"]
            pd = g["provider"]
            daily[(norm_name(cd), dk)]["cust_rev"] += g["rev"]
            daily[(norm_name(cd), dk)]["cust_profit"] += g["profit"]
            daily[(norm_name(cd), dk)]["cust_dur"] += g["dur"]
            daily[(norm_name(cd), dk)]["cust_calls"] += 0
            daily[(norm_name(pd), dk)]["prov_exp"] += g["exp"]
            daily[(norm_name(pd), dk)]["prov_profit"] -= g["exp"]
            daily[(norm_name(pd), dk)]["prov_dur"] += g["dur"]
            daily[(norm_name(pd), dk)]["prov_calls"] += 0

    # Display-name maps for normalized customer/provider/destination keys.
    display = {}
    for r in fi_rows:
        for col in ("Customer", "Provider", "Destination"):
            v = str(first(r, col) or "").strip()
            if v:
                display[(col, norm_name(v))] = v
    for g in gp:
        display.setdefault(("Customer", norm_name(g["customer"])), g["customer"])
        display.setdefault(("Provider", norm_name(g["provider"])), g["provider"])
    display[("Destination", norm_name("(unspecified)"))] = "(unspecified)"

    customer_routes = []
    route_agg = defaultdict(lambda: {"sell": 0.0, "buy": 0.0, "profit": 0.0, "dur": 0.0, "calls": 0.0, "asr": 0.0, "acd": 0.0, "n": 0})
    for (ck, dk, pk), v in cr.items():
        customer = display.get(("Customer", ck), ck or "(unspecified)")
        provider = display.get(("Provider", pk), pk or "(unspecified)")
        dest = display.get(("Destination", dk), dk or "(unspecified)")
        sell, buy, profit = v["rev"], v["exp"], v["profit"]
        metrics = fi.get((ck, pk, dk), {})
        asr, acd = metrics.get("asr", 0.0), metrics.get("acd", 0.0)
        customer_routes.append([None, customer, dest, provider, sell, buy, profit, (profit / sell * 100 if sell else 0.0), asr, acd, int(round(v["calls"])), v["dur"], sell, buy])
        rk = (dk, pk)
        ra = route_agg[rk]
        for fld, val in (("sell", sell), ("buy", buy), ("profit", profit), ("dur", v["dur"]), ("calls", v["calls"])):
            ra[fld] += val
        ra["asr"] += asr; ra["acd"] += acd; ra["n"] += 1

    # Keep customer_routes ordering deterministic, then assign contiguous IDs.
    customer_routes.sort(key=lambda x: (x[1].lower(), x[2].lower(), x[3].lower()))
    for i, row in enumerate(customer_routes):
        row[0] = i

    routes = []
    for (dk, pk), v in route_agg.items():
        dest = display.get(("Destination", dk), dk or "(unspecified)")
        provider = display.get(("Provider", pk), pk or "(unspecified)")
        hits = lcr.get((dk, pk), [])
        trunk = hits[0][2] if hits else None
        rate = min((x[3] for x in hits), default=None)
        routes.append([None, dest, provider, trunk, rate, v["buy"], v["sell"], v["profit"], (v["profit"] / v["sell"] * 100 if v["sell"] else 0.0), (v["asr"] / v["n"] if v["n"] else 0.0), (v["acd"] / v["n"] if v["n"] else 0.0), int(round(v["calls"])), v["dur"]])
    routes.sort(key=lambda x: (x[1].lower(), x[2].lower()))
    for i, row in enumerate(routes):
        row[0] = i

    # Exposure and carrier-details data.
    exposure_carrier = find_col(ex_h, exact=("Carrier", "Carrier Name"), contains=("carrier",))
    current_exp = find_col(ex_h, exact=("Current Exposure",))
    credit = find_col(ex_h, exact=("Credit Limit (Customer)",), contains=("credit", "limit"))
    due_col = find_col(ex_h, exact=("Due", "Due Date", "Payment Due"), contains=("due",))
    netbal_col = find_col(ex_h, exact=("Net Balance", "Netbal"), contains=("net", "balance"))
    terms_col = find_col(ex_h, exact=("Terms", "Payment Terms"), contains=("terms",))
    conf_col = find_col(ex_h, exact=("Confidence", "Conf"), contains=("conf",))
    am_display_col = find_col(ex_h, exact=("Account Manager", "AM", "Account Manager Name"), contains=("account", "manager"))

    exposure = {}
    for r in ex_rows:
        name = str(r.get(exposure_carrier) or "").strip() if exposure_carrier else ""
        if not name:
            continue
        exposure[norm_name(name)] = {
            "name": name,
            "exposure": to_float(r.get(current_exp)),
            "credit": to_float(r.get(credit)),
            "due": str(r.get(due_col) or "").strip() if due_col else None,
            "netbal": to_float(r.get(netbal_col)),
            "terms": str(r.get(terms_col) or "").strip() if terms_col else None,
            "conf": str(r.get(conf_col) or "").strip() if conf_col else None,
            "am_display": str(r.get(am_display_col) or "").strip() if am_display_col else "",
        }

    username_to_display = {}
    cd_carrier = find_col(cd_h, exact=("Carrier Name",))
    cd_am = find_col(cd_h, exact=("Carrier Account Manager",))
    for r in cd_rows:
        cname = str(r.get(cd_carrier) or "").strip() if cd_carrier else ""
        raw_user = str(r.get(cd_am) or "").strip() if cd_am else ""
        ex = exposure.get(norm_name(cname))
        if not ex or not raw_user:
            continue
        users = [x.strip() for x in re.split(r"[,;]+", raw_user) if x.strip()]
        displays = [x.strip() for x in re.split(r"[,;]+", ex.get("am_display", "")) if x.strip()]
        for i, user in enumerate(users):
            if i < len(displays):
                username_to_display[user] = displays[i]
        if users and displays and len(users) == 1:
            username_to_display[users[0]] = displays[0]

    # Identity universe comes from GP + Full info + Exposure + Carrier Details.
    names = {}
    for r in ex_rows:
        n = str(r.get(exposure_carrier) or "").strip() if exposure_carrier else ""
        if n: names.setdefault(norm_name(n), n)
    for r in gp:
        if r["customer"]: names.setdefault(norm_name(r["customer"]), r["customer"])
        if r["provider"]: names.setdefault(norm_name(r["provider"]), r["provider"])
    for r in fi_rows:
        for c in ("Customer", "Provider"):
            n = str(first(r, c) or "").strip()
            if n: names.setdefault(norm_name(n), n)
    for r in cd_rows:
        n = str(r.get(cd_carrier) or "").strip() if cd_carrier else ""
        if n: names.setdefault(norm_name(n), n)

    cust_agg = defaultdict(lambda: {"dur":0.0,"calls":0,"rev":0.0,"profit":0.0})
    prov_agg = defaultdict(lambda: {"dur":0.0,"calls":0,"exp":0.0,"profit":0.0})
    for g in gp:
        ck, pk = norm_name(g["customer"]), norm_name(g["provider"])
        cust_agg[ck]["dur"] += g["dur"]; cust_agg[ck]["rev"] += g["rev"]; cust_agg[ck]["profit"] += g["profit"]
        prov_agg[pk]["dur"] += g["dur"]; prov_agg[pk]["exp"] += g["exp"]; prov_agg[pk]["profit"] -= g["exp"]
    # Calls are not present in Gross Profit; Full info supplies route calls.
    for (ck, dk, pk), v in cr.items():
        cust_agg[ck]["calls"] += int(round(v["calls"])); prov_agg[pk]["calls"] += int(round(v["calls"]))

    roles = {}
    for k in names:
        c, p = k in cust_agg, k in prov_agg
        roles[k] = "both" if c and p else "customer" if c else "provider" if p else "customer"

    def am_for(k):
        ex = exposure.get(k, {})
        return ex.get("am_display") or "Unassigned"

    carriers = []
    for k, name in sorted(names.items(), key=lambda x: x[1].lower()):
        c, p = cust_agg[k], prov_agg[k]
        ex = exposure.get(k, {})
        carriers.append([name, roles[k], am_for(k), c["dur"] + p["dur"], c["calls"] + p["calls"], c["rev"], c["profit"] + p["profit"], ex.get("due"), ex.get("netbal"), ex.get("exposure"), p["exp"]])

    def am_breakdowns():
        groups = defaultdict(lambda: {"profit":0.0,"calls":0,"dur":0.0,"rev":0.0,"exp":0.0,"routes":0})
        for k, a in cust_agg.items():
            am = am_for(k); g = groups[(am, "c", names.get(k, k))]
            g["profit"] += a["profit"]; g["calls"] += a["calls"]; g["dur"] += a["dur"]; g["rev"] += a["rev"]; g["routes"] += sum(1 for x in customer_routes if norm_name(x[1]) == k)
        for k, a in prov_agg.items():
            am = am_for(k); g = groups[(am, "p", names.get(k, k))]
            g["profit"] += a["profit"]; g["calls"] += a["calls"]; g["dur"] += a["dur"]; g["exp"] += a["exp"]; g["routes"] += sum(1 for x in routes if norm_name(x[2]) == k)
        out = []
        for (am, side, name), g in sorted(groups.items()):
            out.append([None, am, side, name, g["profit"], g["calls"], g["dur"], g["rev"], g["exp"], g["routes"], int(round(g["dur"] * 60))])
        for i, r in enumerate(out): r[0] = i
        return out

    am_rows = am_breakdowns()
    touched = defaultdict(set)
    for k in set(cust_agg) | set(prov_agg): touched[am_for(k)].add(names.get(k, k))
    am_totals = [[am, len(vals)] for am, vals in sorted(touched.items())]

    daily_rows = []
    # Rebuild daily metrics by carrier with the exact GP monetary/duration source.
    dd = defaultdict(lambda: {"cust_rev":0.0,"cust_profit":0.0,"cust_dur":0.0,"cust_calls":0,"prov_exp":0.0,"prov_profit":0.0,"prov_dur":0.0,"prov_calls":0})
    for g in gp:
        if not g["date"]: continue
        ck, pk = norm_name(g["customer"]), norm_name(g["provider"])
        dd[(ck,g["date"])]["cust_rev"] += g["rev"]; dd[(ck,g["date"])]["cust_profit"] += g["profit"]; dd[(ck,g["date"])]["cust_dur"] += g["dur"]
        dd[(pk,g["date"])]["prov_exp"] += g["exp"]; dd[(pk,g["date"])]["prov_profit"] -= g["exp"]; dd[(pk,g["date"])]["prov_dur"] += g["dur"]
    for (k, day), v in sorted(dd.items(), key=lambda x: (x[0][0], x[0][1])):
        daily_rows.append([None, names.get(k,k), day, v["cust_rev"], v["cust_profit"], v["cust_dur"], v["cust_calls"], v["prov_exp"], v["prov_profit"], v["prov_dur"], v["prov_calls"]])
    for i, r in enumerate(daily_rows): r[0] = i

    outputs = {
        "providers": [], "customers": [], "carriers": carriers, "routes": routes,
        "customer_routes": customer_routes, "am_breakdown": am_rows, "am_totals": am_totals, "daily_carrier": daily_rows,
    }
    # Provider/customer reference rows. Every identity is included in its role table.
    for k, name in sorted(names.items(), key=lambda x: x[1].lower()):
        ex = exposure.get(k, {})
        am = am_for(k)
        base = [name, am, ex.get("exposure"), ex.get("credit"), ex.get("terms"), ex.get("conf")]
        if k in prov_agg: outputs["providers"].append(base)
        if k in cust_agg: outputs["customers"].append(base)

    def fmt(v):
        if v is None: return ""
        if isinstance(v, float):
            if not math.isfinite(v): return ""
            return f"{v:.12f}".rstrip("0").rstrip(".")
        return str(v)

    for table, rows in outputs.items():
        path = args.out / f"{table}.csv"
        with open(path, "w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh, lineterminator="\n")
            w.writerow(SCHEMA[table])
            w.writerows([[fmt(x) for x in r] for r in rows])

    actual = {"rows": {k: len(v) for k,v in outputs.items()}, **gp_total}
    print("Weekly build complete")
    print("Gross Profit totals (grand-total row excluded):")
    for k in ("rev", "exp", "profit", "dur"): print(f"  {k}: {gp_total[k]:.6f}")
    print("Output rows:")
    for k in SCHEMA: print(f"  {k}: {len(outputs[k])}")
    print(f"routes ids contiguous: {all(r[0] == i for i,r in enumerate(routes))}")
    print(f"am_breakdown sides valid: {all(r[2] in {'c','p'} for r in am_rows)}")

    # Hard integrity checks requested by the weekly process.
    for k in ("rev", "exp", "profit", "dur"):
        got = sum(float(r[12] if k=='rev' else r[13] if k=='exp' else r[6] if k=='profit' else r[11]) for r in customer_routes)
        if abs(got - gp_total[k]) > 1e-6:
            raise SystemExit(f"Integrity failure: customer_routes {k}={got:.12f}, GP={gp_total[k]:.12f}")
    if not all(r[0] == i for i,r in enumerate(routes)):
        raise SystemExit("Integrity failure: routes.id is not contiguous from 0")
    if not all(r[2] in {"c", "p"} for r in am_rows):
        raise SystemExit("Integrity failure: am_breakdown.side contains an invalid value")


if __name__ == "__main__":
    main()
